"""
This script processes employee and facility data to generate an output Excel file
that provides information about the closest facilities to each employee's location.

Input files:
- input.csv: Contains employee name and zip code (columns "Name" and "Employee Zip")
- facilities.csv: Contains facility zip code and airport code (columns "Facility Zip" and "Airport Code")

Output file:
- output.xlsx: Contains the 3 closest facilities for each employee, including the facility zip code, distance in miles, and airport code.
"""

import asyncio
import aiohttp
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment

API_KEY = '***************************************'
INPUT_FILE = 'input.csv'
FACILITIES_FILE = 'facilities.csv'
OUTPUT_FILE = 'output.xlsx'


async def test_network_connectivity():
    """
    Checks the network connectivity by attempting to connect to Google.com.

    Returns:
        bool: True if the network connectivity test is successful, False otherwise.
    """
    async with aiohttp.ClientSession() as session:
        try:
            async with session.get('http://www.google.com') as response:
                return response.status == 200
        except aiohttp.ClientConnectionError:
            return False


async def test_api_connection(api_key):
    """
    Checks the connection to the Google Maps Distance Matrix API.

    Args:
        api_key (str): The Google Maps Distance Matrix API key.

    Returns:
        bool: True if the API connection test is successful, False otherwise.
    """
    async with aiohttp.ClientSession() as session:
        try:
            async with session.get(
                    f'https://maps.googleapis.com/maps/api/distancematrix/json?origins=New+York,NY&destinations=Los+Angeles,CA&key={api_key}') as response:
                data = await response.json()
                return data['status'] == 'OK'
        except aiohttp.ClientConnectionError:
            return False


async def process_distances_batch(api_key, employee_zip, facility_zips, session):
    """
    Processes the distances between the employee zip code and the facility zip codes in batches.

    Args:
        api_key (str): The Google Maps Distance Matrix API key.
        employee_zip (str): The employee zip code.
        facility_zips (list): A list of facility zip codes.
        session (aiohttp.ClientSession): The aiohttp client session.

    Returns:
        list: A list of tuples, where each tuple contains the facility zip code and the distance text.
    """
    batch_size = 10
    facility_batches = [facility_zips[i:i + batch_size] for i in range(0, len(facility_zips), batch_size)]
    facility_batches = [[str(zip) for zip in batch] for batch in
                        facility_batches]  # Convert facility zip codes to strings

    distances = []
    for batch in facility_batches:
        destinations = '|'.join(batch)
        elements = await fetch_distance(session, employee_zip, destinations)
        distances.extend(elements)

    result = []
    for facility_zip, distance in zip(facility_zips, distances):
        facility_zip = str(facility_zip)  # Convert to string to preserve leading zeros
        if 'distance' in distance:
            result.append((facility_zip, distance['distance']['text']))
        else:
            print(f"Distance data not available for facility {facility_zip}.")
    return result


def km_to_miles(km):
    """
    Converts kilometers to miles.

    Args:
        km (float): The distance in kilometers.

    Returns:
        float: The distance in miles.
    """
    return 0.621371 * km


def generate_output_file(data, facility_codes_mapping, employee_names, output_file):
    """
    Generates the output Excel file with the processed data.

    Args:
        data (list): A list of tuples, where each tuple contains the employee zip code and a list of the 3 closest facility zip codes and distances.
        facility_codes_mapping (dict): A dictionary mapping facility zip codes to airport codes.
        employee_names (list): A list of employee names.
        output_file (str): The name of the output Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws['A1'] = "Employee Name"
    ws['B1'] = "Employee Zip"
    for i in range(3):
        ws.cell(row=1, column=i + 3, value=f"Closest Facility {i + 1}")
        ws.cell(row=1, column=i + 6, value=f"Distance {i + 1} (miles)")
        ws.cell(row=1, column=i + 9, value=f"Facility {i + 1} Airport Code")

    for idx, (employee_zip, facilities) in enumerate(data, start=2):
        employee_name = employee_names[idx - 2]
        ws.cell(row=idx, column=1, value=employee_name)
        ws.cell(row=idx, column=2, value=employee_zip)
        for i, (facility_zip, distance) in enumerate(facilities, start=0):
            facility_zip = str(facility_zip)
            ws.cell(row=idx, column=i + 3, value=facility_zip)
            distance_km = float(distance.split()[0].replace(',', ''))
            distance_miles = int(km_to_miles(distance_km))  # Convert to whole number miles
            ws.cell(row=idx, column=i + 6, value=distance_miles)
            facility_airport_code = facility_codes_mapping[facility_zip]
            ws.cell(row=idx, column=i + 9, value=facility_airport_code)

    wb.save(output_file)


async def main():
    """
    The main function that orchestrates the script's execution.
    """
    print("Version 1.0")
    print("Testing network connectivity...")
    network_test = await test_network_connectivity()
    if not network_test:
        print("Network connectivity test failed. Exiting.")
        return
    print("Network connectivity test successful.")

    print("Testing API connection...")
    api_test = await test_api_connection(API_KEY)
    if not api_test:
        print("API connection test failed. Exiting.")
        return
    print("API connection test successful.")

    df_input = pd.read_csv(INPUT_FILE, dtype={'Employee Zip': str})
    employee_names = df_input['Name'].tolist()  # Fetch the "Name" column

    df_facilities = pd.read_csv(FACILITIES_FILE, dtype={'Facility Zip': str})

    facility_codes_mapping = {}
    for _, row in df_facilities.iterrows():
        facility_codes_mapping[str(row['Facility Zip'])] = row['Airport Code']

    async with aiohttp.ClientSession() as session:
        data = []
        for idx, row in df_input.iterrows():
            employee_zip = str(row['Employee Zip']).split('.')[0]  # Remove decimal point
            print(f"Processing distances for employee with zip {employee_zip}...")
            facility_zips = list(facility_codes_mapping.keys())
            facility_distances = await process_distances_batch(API_KEY, employee_zip, facility_zips, session)
            facility_distances.sort(key=lambda x: float(x[1].replace(',', '').split()[0]))
            data.append((employee_zip, facility_distances[:3]))
            print(f"Distances processed for employee with zip {employee_zip}.")

        print("Generating output file...")
        generate_output_file(data, facility_codes_mapping, employee_names, OUTPUT_FILE)
        print(f"Results have been saved to {OUTPUT_FILE}.")


if __name__ == "__main__":
    loop = asyncio.get_event_loop()
    loop.run_until_complete(main())